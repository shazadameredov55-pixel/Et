"""
XLSX product generator. Implements src.core.interfaces.ProductGenerator.

Consumes a (ProductSpec, DesignProfile) pair plus the ProductBlueprint
resolved from ProductSpec.niche, and produces a real, working Excel
workbook: live formulas (not baked-in values), number formats, data
validation dropdowns, freeze panes, conditional formatting, and named
ranges the formulas rely on.

Demo/sample rows are seeded ONLY to prove the formulas compute correctly
(see _seed_sample_data + tests/test_financial_formulas.py), then either
removed or clearly marked "DEMO" before the file is considered final —
controlled by `clear_demo_data`.
"""

from __future__ import annotations

import logging
import os
from typing import Any

from openpyxl import Workbook
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import CellIsRule, FormulaRule, ColorScaleRule
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

from src.core.interfaces import ProductGenerator, GenerationResult, GeneratedFile
from src.core.models import ProductSpec, DesignProfile
from src.generators.blueprints import get_blueprint, ProductBlueprint, SheetSpec

logger = logging.getLogger(__name__)


class XlsxIntegrityError(Exception):
    """Raised when a generated workbook fails a basic integrity check
    (e.g. cannot be reopened, missing an expected sheet)."""


class XlsxGenerator(ProductGenerator):
    output_type = "xlsx"

    def __init__(self, clear_demo_data: bool = True):
        # clear_demo_data=False is used internally by the "financial
        # testing" step (madde 8): seed sample data, verify computed
        # values, THEN generate the real, clean file with it True.
        self.clear_demo_data = clear_demo_data

    # ------------------------------------------------------------------
    # Public entrypoint (ProductGenerator contract)
    # ------------------------------------------------------------------

    def generate(self, product_spec: ProductSpec, design_profile: DesignProfile, output_dir: str) -> GenerationResult:
        try:
            blueprint = get_blueprint(product_spec.niche)
        except KeyError as e:
            logger.error("XLSX generation failed: %s", e)
            return GenerationResult(success=False, files=[], error=str(e))

        try:
            wb = Workbook()
            # Remove the default sheet openpyxl creates; we add our own in order.
            default_sheet = wb.active
            wb.remove(default_sheet)

            style = _StyleKit(design_profile)

            sheet_objs: dict[str, Worksheet] = {}
            for sheet_spec in blueprint.sheets:
                ws = wb.create_sheet(title=sheet_spec.title[:31])  # Excel sheet name limit
                sheet_objs[sheet_spec.sheet_key] = ws
                self._build_sheet(ws, sheet_spec, style)

            self._seed_sample_data(sheet_objs, blueprint)
            self._apply_formulas(sheet_objs, blueprint, product_spec)
            self._write_instructions(sheet_objs.get("instructions"), product_spec, blueprint, style)

            if self.clear_demo_data:
                self._clear_demo_rows(sheet_objs, blueprint)

            os.makedirs(output_dir, exist_ok=True)
            file_path = os.path.join(output_dir, "product.xlsx")
            wb.save(file_path)

        except Exception as e:  # noqa: BLE001 - must not raise past this boundary
            logger.exception("XLSX generation raised an unexpected error")
            return GenerationResult(success=False, files=[], error=str(e))

        try:
            self._verify_integrity(file_path, blueprint)
        except XlsxIntegrityError as e:
            logger.error("XLSX integrity check failed: %s", e)
            return GenerationResult(success=False, files=[], error=str(e))

        return GenerationResult(
            success=True,
            files=[
                GeneratedFile(
                    file_path=file_path,
                    file_type="xlsx",
                    description=f"{blueprint.display_name} workbook",
                )
            ],
        )

    # ------------------------------------------------------------------
    # Sheet construction
    # ------------------------------------------------------------------

    def _build_sheet(self, ws: Worksheet, spec: SheetSpec, style: "_StyleKit") -> None:
        # Header row
        for col_idx, col in enumerate(spec.columns, start=1):
            cell = ws.cell(row=1, column=col_idx, value=col.header)
            cell.font = style.header_font
            cell.fill = style.header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = style.thin_border
            ws.column_dimensions[get_column_letter(col_idx)].width = col.width

        ws.row_dimensions[1].height = max(20, style.row_height)
        ws.freeze_panes = spec.freeze_panes

        # Number formats + data validation applied down a reasonable
        # working range (sample rows + headroom for the user to add more).
        working_rows = max(spec.sample_rows, 6) + 44  # generous headroom, still bounded
        for col_idx, col in enumerate(spec.columns, start=1):
            col_letter = get_column_letter(col_idx)
            if col.number_format:
                for row in range(spec.starting_data_row, spec.starting_data_row + working_rows):
                    ws.cell(row=row, column=col_idx).number_format = col.number_format
            if col.data_validation:
                dv = DataValidation(
                    type=col.data_validation["type"],
                    formula1=col.data_validation["formula1"],
                    allow_blank=True,
                    showDropDown=False,  # False = show the dropdown arrow (openpyxl quirk)
                )
                dv.error = "Please choose a value from the list."
                dv.errorTitle = "Invalid entry"
                rng = f"{col_letter}{spec.starting_data_row}:{col_letter}{spec.starting_data_row + working_rows}"
                dv.add(rng)
                ws.add_data_validation(dv)

        if spec.has_totals_row:
            total_row = spec.starting_data_row + working_rows + 1
            label_cell = ws.cell(row=total_row, column=1, value="Total")
            label_cell.font = Font(bold=True)
            for col_idx, col in enumerate(spec.columns, start=1):
                if col.number_format and col_idx > 1:
                    col_letter = get_column_letter(col_idx)
                    first = spec.starting_data_row
                    last = spec.starting_data_row + working_rows - 1
                    total_cell = ws.cell(
                        row=total_row, column=col_idx,
                        value=f"=SUM({col_letter}{first}:{col_letter}{last})",
                    )
                    total_cell.number_format = col.number_format
                    total_cell.font = Font(bold=True)

        for rule_id in spec.conditional_formatting:
            self._apply_conditional_formatting(ws, spec, rule_id, working_rows)

    def _apply_conditional_formatting(self, ws: Worksheet, spec: SheetSpec, rule_id: str, working_rows: int) -> None:
        first = spec.starting_data_row
        last = spec.starting_data_row + working_rows - 1

        red_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
        green_fill = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")

        # Find the "Amount"-like column for value-based rules.
        amount_col_idx = next(
            (i for i, c in enumerate(spec.columns, start=1) if "Amount" in c.header or "Balance" in c.header),
            None,
        )

        if rule_id == "highlight_negative" and amount_col_idx:
            col_letter = get_column_letter(amount_col_idx)
            rng = f"{col_letter}{first}:{col_letter}{last}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="lessThan", formula=["0"], fill=red_fill)
            )
        elif rule_id == "highlight_over_budget" and amount_col_idx:
            col_letter = get_column_letter(amount_col_idx)
            rng = f"{col_letter}{first}:{col_letter}{last}"
            ws.conditional_formatting.add(
                rng, CellIsRule(operator="greaterThan", formula=["500"], fill=red_fill)
            )
        elif rule_id == "highlight_overdue":
            status_col_idx = next((i for i, c in enumerate(spec.columns) if "Status" in spec.columns[i].header), None)
            if status_col_idx:
                col_letter = get_column_letter(status_col_idx + 1)
                rng = f"{col_letter}{first}:{col_letter}{last}"
                ws.conditional_formatting.add(
                    rng, FormulaRule(formula=[f'{col_letter}{first}="Overdue"'], fill=red_fill)
                )
        elif rule_id == "progress_bar":
            pct_col_idx = next((i for i, c in enumerate(spec.columns, start=1) if "%" in c.header), None)
            if pct_col_idx:
                col_letter = get_column_letter(pct_col_idx)
                rng = f"{col_letter}2:{col_letter}20"
                ws.conditional_formatting.add(
                    rng,
                    ColorScaleRule(
                        start_type="min", start_color="FCA5A5",
                        end_type="max", end_color="86EFAC",
                    ),
                )

    # ------------------------------------------------------------------
    # Sample data (for financial verification), formulas, cleanup
    # ------------------------------------------------------------------

    def _seed_sample_data(self, sheets: dict[str, Worksheet], blueprint: ProductBlueprint) -> None:
        """Seed deterministic demo rows. Values chosen so downstream
        formula tests have a known-correct expected result (see
        tests/test_financial_formulas.py)."""
        if "income" in sheets:
            ws = sheets["income"]
            demo = [("2026-01-01", "Salary", 3500.00), ("2026-01-15", "Freelance", 1500.00)]
            for i, (date, source, amount) in enumerate(demo, start=2):
                ws.cell(row=i, column=1, value=date)
                ws.cell(row=i, column=2, value=f"DEMO: {source}")
                ws.cell(row=i, column=3, value=amount)

        if "expenses" in sheets:
            ws = sheets["expenses"]
            demo = [
                ("2026-01-02", "Housing", "DEMO: Rent", 1800.00),
                ("2026-01-05", "Food", "DEMO: Groceries", 700.00),
                ("2026-01-10", "Transport", "DEMO: Gas", 500.00),
            ]
            for i, (date, cat, desc, amount) in enumerate(demo, start=2):
                ws.cell(row=i, column=1, value=date)
                ws.cell(row=i, column=2, value=cat)
                ws.cell(row=i, column=3, value=desc)
                ws.cell(row=i, column=4, value=amount)

        if "client_income" in sheets:
            ws = sheets["client_income"]
            demo = [
                ("2026-01-05", "DEMO: Client A", "Paid", 2000.00),
                ("2026-01-20", "DEMO: Client B", "Unpaid", 1200.00),
            ]
            for i, (date, client, status, amount) in enumerate(demo, start=2):
                ws.cell(row=i, column=1, value=date)
                ws.cell(row=i, column=2, value=client)
                ws.cell(row=i, column=3, value=status)
                ws.cell(row=i, column=4, value=amount)

        if "debts" in sheets:
            ws = sheets["debts"]
            demo = [("DEMO: Credit Card A", "Credit Card", 2000.00, 0.199, 50.00)]
            for i, (name, dtype, bal, rate, minp) in enumerate(demo, start=2):
                ws.cell(row=i, column=1, value=name)
                ws.cell(row=i, column=2, value=dtype)
                ws.cell(row=i, column=3, value=bal)
                ws.cell(row=i, column=4, value=rate)
                ws.cell(row=i, column=5, value=minp)

        if "payment_log" in sheets:
            ws = sheets["payment_log"]
            demo = [("2026-01-10", "DEMO: Credit Card A", 500.00)]
            for i, (date, name, amount) in enumerate(demo, start=2):
                ws.cell(row=i, column=1, value=date)
                ws.cell(row=i, column=2, value=name)
                ws.cell(row=i, column=3, value=amount)

    def _apply_formulas(self, sheets: dict[str, Worksheet], blueprint: ProductBlueprint, spec: ProductSpec) -> None:
        """Wire up named ranges + cross-sheet formulas so the dashboard
        reflects live totals rather than baked-in numbers."""
        wb = next(iter(sheets.values())).parent

        if blueprint.niche == "monthly_budget_tracker":
            income_ws, expenses_ws, dash_ws = sheets["income"], sheets["expenses"], sheets["dashboard"]
            self._define_name(wb, "Income_Total", f"'{income_ws.title}'!$C$2:$C$50")
            self._define_name(wb, "Expenses_Total", f"'{expenses_ws.title}'!$D$2:$D$50")
            dash_ws["A1"] = "Total Income"
            dash_ws["B1"] = "=SUM(Income_Total)"
            dash_ws["B1"].number_format = "#,##0.00"
            dash_ws["A2"] = "Total Expenses"
            dash_ws["B2"] = "=SUM(Expenses_Total)"
            dash_ws["B2"].number_format = "#,##0.00"
            dash_ws["A3"] = "Remaining"
            dash_ws["B3"] = "=B1-B2"
            dash_ws["B3"].number_format = "#,##0.00"

        elif blueprint.niche == "freelancer_finance_tracker":
            income_ws = sheets["client_income"]
            exp_ws = sheets["business_expenses"]
            dash_ws = sheets["dashboard"]
            self._define_name(wb, "Gross_Income", f"'{income_ws.title}'!$D$2:$D$50")
            self._define_name(wb, "Business_Expenses_Total", f"'{exp_ws.title}'!$D$2:$D$50")
            dash_ws["A1"] = "Gross Income (YTD)"
            dash_ws["B1"] = "=SUM(Gross_Income)"
            dash_ws["A2"] = "Business Expenses (YTD)"
            dash_ws["B2"] = "=SUM(Business_Expenses_Total)"
            dash_ws["A3"] = "Net Income (YTD)"
            dash_ws["B3"] = "=B1-B2"
            for col in ("B1", "B2", "B3"):
                dash_ws[col].number_format = "#,##0.00"

            tax_ws = sheets["tax_setaside"]
            for row in range(2, 6):
                # Set-Aside Amount = Net Income * Tax Rate. Blank inputs
                # multiply to 0, which is harmless until the user fills them in.
                cell = tax_ws.cell(row=row, column=4, value=f"=B{row}*C{row}")
                cell.number_format = "#,##0.00"

            dash_ws["A4"] = "Outstanding Invoices"
            dash_ws["B4"] = f"=SUMIF('{income_ws.title}'!C2:C50,\"<>Paid\",'{income_ws.title}'!D2:D50)"
            dash_ws["B4"].number_format = "#,##0.00"

        elif blueprint.niche == "debt_payoff_tracker":
            debts_ws = sheets["debts"]
            payments_ws = sheets["payment_log"]
            dash_ws = sheets["dashboard"]
            # One dashboard row per debt (matched by name); simple MVP
            # approach — up to 10 debts.
            for i in range(2, 12):
                debt_name_cell = f"'{debts_ws.title}'!A{i}"
                dash_ws.cell(row=i, column=1, value=f"={debt_name_cell}")
                dash_ws.cell(row=i, column=2, value=f"='{debts_ws.title}'!C{i}")
                dash_ws.cell(row=i, column=2).number_format = "#,##0.00"
                remaining_formula = (
                    f"=MAX('{debts_ws.title}'!C{i}-SUMIF('{payments_ws.title}'!B:B,"
                    f"'{debts_ws.title}'!A{i},'{payments_ws.title}'!C:C),0)"
                )
                dash_ws.cell(row=i, column=3, value=remaining_formula)
                dash_ws.cell(row=i, column=3).number_format = "#,##0.00"
                progress_formula = (
                    f"=IF('{debts_ws.title}'!C{i}=0,0,1-(C{i}/'{debts_ws.title}'!C{i}))"
                )
                dash_ws.cell(row=i, column=4, value=progress_formula)
                dash_ws.cell(row=i, column=4).number_format = "0.0%"

    def _write_instructions(
        self, ws: Worksheet | None, product_spec: ProductSpec, blueprint: ProductBlueprint, style: "_StyleKit"
    ) -> None:
        if ws is None:
            return
        lines = [
            f"{blueprint.display_name} — How to use this workbook",
            "",
            f"Who this is for: {product_spec.target_customer or blueprint.default_target_customer}",
            "",
            "Getting started:",
            "1. Enter your own data in the input sheets (rows marked DEMO can be deleted).",
            "2. Totals and percentages update automatically — you do not need to edit formula cells.",
            "3. Use the dropdown lists provided in category/status columns for consistent data entry.",
            "",
            "This template is provided for personal budgeting organization only and does not",
            "constitute financial advice.",
        ]
        for i, line in enumerate(lines, start=1):
            cell = ws.cell(row=i, column=1, value=line)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
        ws.column_dimensions["A"].width = 90

    def _clear_demo_rows(self, sheets: dict[str, Worksheet], blueprint: ProductBlueprint) -> None:
        """Remove seeded demo values from data-entry cells so the shipped
        product starts empty for the buyer. Formula cells and headers are
        left untouched."""
        data_sheet_keys = {
            "income", "expenses", "client_income", "business_expenses",
            "tax_setaside", "debts", "payment_log",
        }
        for key in data_sheet_keys & sheets.keys():
            ws = sheets[key]
            for row in ws.iter_rows(min_row=2, max_row=15):
                for cell in row:
                    if cell.value is not None and not (isinstance(cell.value, str) and cell.value.startswith("=")):
                        cell.value = None

    # ------------------------------------------------------------------
    # Integrity verification
    # ------------------------------------------------------------------

    def _verify_integrity(self, file_path: str, blueprint: ProductBlueprint) -> None:
        """Reopen the saved file and check it is well-formed: opens
        without error, has every expected sheet, and every sheet has a
        header row. This is a structural check, not a full formula
        evaluation (openpyxl does not evaluate formulas)."""
        from openpyxl import load_workbook

        try:
            wb = load_workbook(file_path, data_only=False)
        except Exception as e:
            raise XlsxIntegrityError(f"Workbook could not be reopened: {e}") from e

        expected_titles = {s.title[:31] for s in blueprint.sheets}
        actual_titles = set(wb.sheetnames)
        missing = expected_titles - actual_titles
        if missing:
            raise XlsxIntegrityError(f"Missing expected sheets: {missing}")

        for sheet_spec in blueprint.sheets:
            ws = wb[sheet_spec.title[:31]]
            header_values = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
            if not any(header_values):
                raise XlsxIntegrityError(f"Sheet '{sheet_spec.title}' has an empty header row")

    def _define_name(self, wb, name: str, ref: str) -> None:
        if name in wb.defined_names:
            del wb.defined_names[name]
        wb.defined_names[name] = DefinedName(name, attr_text=ref)


class _StyleKit:
    """Resolves a DesignProfile into concrete openpyxl style objects."""

    def __init__(self, design_profile: DesignProfile):
        typography = design_profile.typography or {}
        table_design = design_profile.table_design or {}
        spacing = design_profile.spacing or {}

        heading_font = typography.get("heading_font", "Calibri")
        header_fill_hex = table_design.get("header_fill", "1F2937")
        header_font_color = table_design.get("header_font_color", "FFFFFF")

        self.header_font = Font(name=heading_font, bold=True, color=header_font_color, size=11)
        self.header_fill = PatternFill(start_color=header_fill_hex, end_color=header_fill_hex, fill_type="solid")
        self.row_height = spacing.get("row_height", 18)
        thin = Side(style="thin", color="D1D5DB")
        self.thin_border = Border(left=thin, right=thin, top=thin, bottom=thin)
