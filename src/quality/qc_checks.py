"""
Quality control engine (requirement #7 of Phase 2 / #8 of Phase 3).

This module does NOT re-implement the integrity checks already inside
each generator (XlsxGenerator._verify_integrity, etc.) — those catch
"can this even be opened" failures at generation time and already cause
GenerationResult.success=False. QCEngine runs AFTER a successful
generation + packaging, as an independent, holistic pass over the final
packaged files, producing a single 0-100 score with itemized issues that
feed the AI self-critique step and the revision loop.

Design choice: scoring is deterministic and rule-based here (not an LLM
call) so it never depends on a paid API and is exactly reproducible in
tests. self_critique.py adds a qualitative layer on top using AIProvider,
but the pass/fail QUALITY GATE is always grounded in these deterministic
checks first — an LLM's opinion can lower a score further but a
structurally broken file can never pass because an LLM said it looked fine.
"""

from __future__ import annotations

import logging
import os
import zipfile

from openpyxl import load_workbook
from pypdf import PdfReader
from PIL import Image

from src.core.models import QCResult, QCIssue, ProductSpec
from src.generators.blueprints import ProductBlueprint, get_blueprint

logger = logging.getLogger(__name__)

# Point deductions per issue severity, off a 100-point baseline.
_DEDUCTIONS = {"critical": 40, "major": 15, "minor": 5}


class QCEngine:
    def __init__(self, minimum_pass_score: float = 80.0):
        self.minimum_pass_score = minimum_pass_score

    def check_package(self, product_spec: ProductSpec, package_dir: str, zip_path: str) -> QCResult:
        issues: list[QCIssue] = []
        checked_files: list[str] = []

        try:
            blueprint = get_blueprint(product_spec.niche)
        except KeyError as e:
            issues.append(QCIssue("critical", "unknown_niche", str(e)))
            return QCResult(score=0.0, passed=False, issues=issues, checked_files=checked_files)

        issues += self._check_xlsx(os.path.join(package_dir, "product.xlsx"), blueprint, checked_files)
        issues += self._check_pdf(os.path.join(package_dir, "printable.pdf"), checked_files, label="printable.pdf")
        issues += self._check_pdf(os.path.join(package_dir, "instructions.pdf"), checked_files, label="instructions.pdf")
        issues += self._check_image(os.path.join(package_dir, "preview.png"), checked_files)
        issues += self._check_zip(zip_path, checked_files)

        score = self._score_from_issues(issues)
        passed = score >= self.minimum_pass_score and not any(i.severity == "critical" for i in issues)
        return QCResult(score=score, passed=passed, issues=issues, checked_files=checked_files)

    # ------------------------------------------------------------------

    def _score_from_issues(self, issues: list[QCIssue]) -> float:
        score = 100.0
        for issue in issues:
            score -= _DEDUCTIONS.get(issue.severity, 5)
        return round(max(0.0, min(100.0, score)), 1)

    def _check_xlsx(self, path: str, blueprint: ProductBlueprint, checked_files: list[str]) -> list[QCIssue]:
        checked_files.append(path)
        issues: list[QCIssue] = []
        if not os.path.isfile(path):
            return [QCIssue("critical", "xlsx_missing", f"XLSX file not found: {path}")]

        try:
            wb = load_workbook(path, data_only=False)
        except Exception as e:
            return [QCIssue("critical", "xlsx_unopenable", f"Workbook could not be opened: {e}")]

        expected_titles = {s.title[:31] for s in blueprint.sheets}
        actual_titles = set(wb.sheetnames)
        missing_sheets = expected_titles - actual_titles
        if missing_sheets:
            issues.append(QCIssue("critical", "xlsx_missing_sheets", f"Missing sheets: {missing_sheets}"))

        formula_found = False
        broken_reference_found = False
        for sheet_title in actual_titles & expected_titles:
            ws = wb[sheet_title]
            for row in ws.iter_rows():
                for cell in row:
                    if isinstance(cell.value, str) and cell.value.startswith("="):
                        formula_found = True
                        if "#REF!" in cell.value or "#NAME?" in cell.value:
                            broken_reference_found = True

        if not formula_found:
            issues.append(QCIssue("major", "xlsx_no_formulas", "No live formulas found in the workbook"))
        if broken_reference_found:
            issues.append(QCIssue("critical", "xlsx_broken_reference", "A formula contains a broken reference (#REF!/#NAME?)"))

        has_validation = any(len(wb[t].data_validations.dataValidation) > 0 for t in actual_titles if t in wb.sheetnames)
        if not has_validation:
            issues.append(QCIssue("minor", "xlsx_no_data_validation", "No data validation (dropdowns) found"))

        return issues

    def _check_pdf(self, path: str, checked_files: list[str], label: str) -> list[QCIssue]:
        checked_files.append(path)
        if not os.path.isfile(path):
            return [QCIssue("critical", "pdf_missing", f"{label} not found: {path}")]

        try:
            reader = PdfReader(path)
        except Exception as e:
            return [QCIssue("critical", "pdf_unopenable", f"{label} could not be opened: {e}")]

        issues: list[QCIssue] = []
        if len(reader.pages) < 1:
            issues.append(QCIssue("critical", "pdf_zero_pages", f"{label} has zero pages"))
        for i, page in enumerate(reader.pages):
            text = (page.extract_text() or "").strip()
            if not text:
                issues.append(QCIssue("major", "pdf_blank_page", f"{label} page {i + 1} appears blank"))
        return issues

    def _check_image(self, path: str, checked_files: list[str]) -> list[QCIssue]:
        checked_files.append(path)
        if not os.path.isfile(path):
            return [QCIssue("critical", "preview_missing", f"Preview image not found: {path}")]

        try:
            with Image.open(path) as img:
                img.verify()
            with Image.open(path) as img:
                width, height = img.size
                fmt = img.format
        except Exception as e:
            return [QCIssue("critical", "preview_corrupted", f"Preview image could not be verified: {e}")]

        issues: list[QCIssue] = []
        if width < 400 or height < 300:
            issues.append(QCIssue("minor", "preview_too_small", f"Preview dimensions unusually small: {width}x{height}"))
        if fmt != "PNG":
            issues.append(QCIssue("minor", "preview_wrong_format", f"Expected PNG, got {fmt}"))
        return issues

    def _check_zip(self, zip_path: str, checked_files: list[str]) -> list[QCIssue]:
        checked_files.append(zip_path)
        if not os.path.isfile(zip_path):
            return [QCIssue("critical", "zip_missing", f"ZIP not found: {zip_path}")]

        try:
            with zipfile.ZipFile(zip_path, "r") as zf:
                bad_file = zf.testzip()
                if bad_file is not None:
                    return [QCIssue("critical", "zip_corrupted", f"Corrupted ZIP entry: {bad_file}")]
                names = set(zf.namelist())
        except Exception as e:
            return [QCIssue("critical", "zip_unopenable", f"ZIP could not be opened: {e}")]

        required = {"product.xlsx", "printable.pdf", "preview.png", "instructions.pdf", "metadata.json", "README.txt", "listing.txt"}
        missing = required - names
        if missing:
            return [QCIssue("critical", "zip_missing_entries", f"ZIP missing expected files: {missing}")]
        return []
